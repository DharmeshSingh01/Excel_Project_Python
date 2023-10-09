import openpyxl

# inv_file= openpyxl.load_workbook("E:\PythonProject\spreadsheet\inventory.xlsx") 
inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list= inv_file["Sheet1"]
product_per_supplier={}
total_value_per_supplier={}
inventory_under_10={}

# for item in product_list:
#     print(item.value())

for product_row in range(2,product_list.max_row +1):
    supplier_name = product_list.cell(product_row,4).value
    inv = product_list.cell(product_row,2).value
    price = product_list.cell(product_row,3).value
    product_number= product_list.cell(product_row,1).value
    Totalvalue = product_list.cell(product_row,5)

    if supplier_name in product_per_supplier:
        currentvalue = product_per_supplier.get(supplier_name)
        product_per_supplier[supplier_name]=currentvalue+1
    else:
        product_per_supplier[supplier_name]=1

################# Calculate total value of Inventory###############
    if  supplier_name in total_value_per_supplier:
        currentvalue = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name]=currentvalue+ (inv*price)
    else:
        total_value_per_supplier[supplier_name]=(inv*price)
    
#Logic for inventory less than 10
    if inv <10:
        inventory_under_10[int(product_number)]= int(inv)
    
# Total per product
    Totalvalue.value= inv * price


product_per_supplier_row= product_list.cell(76,1)
product_per_supplier_row.value=str(product_per_supplier)

inv_file.save("NewFileWithUpdate1.xlsx")

print(product_per_supplier)
print (total_value_per_supplier)
print (inventory_under_10)




print(product_list.max_row)